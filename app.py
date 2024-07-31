import json
import os
import uuid
import re
import secrets
from datetime import datetime

import boto3
from boto3.dynamodb.conditions import Attr
from botocore.exceptions import NoCredentialsError, BotoCoreError, ClientError

import xlrd
import docx
from docx import Document

from openpyxl import load_workbook

from flask import Flask, render_template, request, jsonify, redirect, url_for, session as login_session
from botocore.exceptions import NoCredentialsError
from botocore.exceptions import ClientError

app = Flask(__name__)
app.secret_key = secrets.token_hex(16)
app.config['UPLOAD_FOLDER'] = 'upload'     #definir carpeta de uploads


MAX_QUESTIONS = 100


ROLE_ARN_WITH_DB_ACCESS = "arn:aws:iam::833148148022:role/RolBaseDeDatos"
ROLE_SESSION_NAME = "ExampleSkillRoleSession"

session = boto3.Session()
sts = session.client('sts')
assumed_role_object = sts.assume_role(
    RoleArn=ROLE_ARN_WITH_DB_ACCESS,
    RoleSessionName=ROLE_SESSION_NAME
)

#Credenciales temporales del rol asumido
credentials = assumed_role_object['Credentials']

#Crea una nueva sesión con las credenciales del rol asumido
assumed_role_session = boto3.Session(
    aws_access_key_id="",
    aws_secret_access_key="",
    region_name='eu-west-1',
)
dynamodb = assumed_role_session.resource('dynamodb')

###################################################################################################################
@app.route('/')
def index():
    return render_template('index.html', login_session=login_session)


@app.route('/examen')
def examen():
    if not login_session.get('user_id'):
        return redirect('/login')
    return render_template('examen.html', login_session=login_session)


@app.route('/cartas', methods=['GET', 'POST'])
def cards():
    if not login_session.get('user_id'):
        return redirect('/login')
    if request.method == 'POST':
        deck_name = request.form.get('deck_name')
        flashcards = obtener_cartas_dynamo(deck_name)
        return jsonify(flashcards) 
    else:
        return render_template('cartas.html', login_session=login_session)


@app.route('/import_cards', methods=['POST'])
def import_cards():
    if check_document_limit(login_session):
        return jsonify({'error': 'Has alcanzado tu límite diario de documentos.'}), 400
    try:
        nombre_mazo = request.form.get('nombre_mazo')
        document = request.files['file']
        table_name = 'CartasParaEstudiar'
        filename = os.path.join(app.config['UPLOAD_FOLDER'], document.filename)
        print(filename)
        document.save(filename)
        table = dynamodb.Table(table_name)
        extracted_data = import_cards_from_txt(filename, nombre_mazo)
        for item in extracted_data:
            table.put_item(Item=item)
        message = f"Extraido y guardado en la base de datos con el nombre de baraja: {nombre_mazo}"
        return jsonify({'message': message})
    except IndexError:
        return jsonify({'error': 'El formato del archivo no cumple con el ejemplo.'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        os.remove(filename)


@app.route('/addcard', methods=['POST'])
def addcard():
    try:
        data = request.get_json()
        questioncard = data.get('questioncard')
        answercard = data.get('answercard')
        deck_name = data.get('deckcards')
        print("Deck name:", deck_name)
        card_id = generate_unique_id()
        table_name = 'CartasParaEstudiar'
        table = dynamodb.Table(table_name)
        item = {
            'card_id': card_id,  
            'deck_name': deck_name,
            'question': questioncard,
            'answer': answercard,
            'user_id': login_session['user_id']
        }
        table.put_item(Item=item)
        addcardmessage = f"Carta añadida a la baraja: {deck_name}"
        return jsonify({'message': addcardmessage})
    except Exception as e:
        return jsonify({'message': 'An error occurred: ' + str(e)})


@app.route('/login', methods=['GET', 'POST'])
def login():
    try:
        if request.method == 'POST':
            table_name = 'Usuarios'
            table = dynamodb.Table(table_name)
            user_id = request.form['id']
            userName = request.form['name']
            response = table.scan(
                FilterExpression=Attr('user_id').eq(user_id) & Attr('userName').eq(userName)
            )
            if response['Items']:
                user_id = response['Items'][0]['user_id']
                userName = response['Items'][0]['userName']
                login_session['user_id'] = user_id
                login_session['userName'] = userName
                return render_template('login.html', user=login_session['userName'], login_session=login_session)
            else:
                return render_template('login.html', error='Id y usuario incorrectos', login_session=login_session)
        return render_template('login.html', login_session=login_session)
    except Exception as e:
        print(e)
        return render_template('login.html', error='Error al interactuar con la base de datos', login_session=login_session)



@app.route('/ajustes')
def ajustes():
    if not login_session.get('user_id'):
        return redirect('/login')
    message = request.args.get('message')
    return render_template('ajustes.html', login_session=login_session, message=message)


@app.route('/actualizar_api_key_chatgpt', methods=['POST'])
def actualizar_api_key_chatgpt():
    api_key_chatgpt = request.form.get('api_key1')

    table_name = 'Usuarios'
    table = dynamodb.Table(table_name)

    if api_key_chatgpt:
        try:
            response = table.update_item(
                Key={
                    'user_id': login_session['user_id']
                },
                UpdateExpression="SET api_key_chatgpt = :gpt",
                ExpressionAttributeValues={
                    ':gpt': api_key_chatgpt
                }
            )
            return redirect(url_for('ajustes', message='API Key ChatGPT actualizada con éxito'))
        except BotoCoreError as e:
            return 'Error al actualizar API Key ChatGPT: {}'.format(e.message)
        except ClientError as e:
            return 'Error al actualizar API Key ChatGPT: {}'.format(e.response['Error']['Message'])


@app.route('/actualizar_api_key_chatpdf', methods=['POST'])
def actualizar_api_key_chatpdf():
    api_key_chatgpt = request.form.get('api_key2')

    table_name = 'Usuarios'
    table = dynamodb.Table(table_name)

    if api_key_chatgpt:
        try:
            response = table.update_item(
                Key={
                    'user_id': login_session['user_id']
                },
                UpdateExpression="SET api_key_chatpdf = :pdf",
                ExpressionAttributeValues={
                    ':pdf': api_key_chatgpt
                }
            )
            return redirect(url_for('ajustes', message='API Key ChatPDF actualizada con éxito'))
        except BotoCoreError as e:
            return 'Error al actualizar API Key ChatGPT: {}'.format(e.message)
        except ClientError as e:
            return 'Error al actualizar API Key ChatGPT: {}'.format(e.response['Error']['Message'])



@app.route('/get_api_key_chatpdf', methods=['GET'])
def get_api_key_chatpdf():
    user_id = login_session['user_id']
    table = dynamodb.Table('Usuarios')

    try:
        response = table.get_item(Key={'user_id': user_id})
        api_key_chatpdf = response['Item']['api_key_chatpdf']
        return jsonify(api_key_chatpdf=api_key_chatpdf)
    except Exception as e:
        return 'Error al obtener API Key ChatPDF: {}'.format(str(e))


@app.route('/get_api_key_chatgpt', methods=['GET'])
def get_api_key_chatgpt():
    user_id = login_session['user_id']
    table = dynamodb.Table('Usuarios')

    try:
        response = table.get_item(Key={'user_id': user_id})
        api_key_chatgpt = response['Item']['api_key_chatgpt']
        return jsonify(api_key_chatgpt=api_key_chatgpt)
    except Exception as e:
        return 'Error al obtener API Key ChatGPT: {}'.format(str(e))


@app.route('/logout')
def logout():
    login_session.pop('user_id', None) 
    login_session.pop('userName', None) 
    return redirect(url_for('index')) 


@app.route('/calificaciones', methods=['GET'])
def get_calificaciones():
    table = dynamodb.Table('Calificaciones')
    response = table.scan(
        FilterExpression=Attr('user_id').eq(login_session['user_id'])
    )
    return jsonify(response['Items'])


@app.route('/calendario')
def mostrar_calendario():
    return render_template('calendario.html', login_session=login_session)

    
@app.route('/eventos', methods=['GET', 'POST'])
def crear_evento():
    try:
        user_id = login_session.get('user_id')
        if not user_id:
            return redirect('/login')
        if request.method == 'GET':
            return render_template('eventos.html', login_session=login_session)
        if request.method == 'POST':
            titulo = request.form['titulo']
            tipo = request.form['tipo']
            fecha_creacion = datetime.now().strftime("%d/%m/%Y")
            fecha_evento = datetime.strptime(request.form['fecha'], "%d/%m/%Y").strftime("%d/%m/%Y")
            descripcion = request.form['descripcion']
            table_name = 'Eventos'
            table = dynamodb.Table(table_name)

            evento = {
                'evento_id': generate_unique_id(),
                'titulo': titulo,
                'tipo': tipo,
                'fecha_creacion': fecha_creacion,
                'fecha_evento': fecha_evento,
                'descripcion': descripcion,
                'user_id': login_session['user_id']
            }
            table.put_item(Item=evento)
            mensaje_exito = "Evento creado con éxito"
            return render_template('eventos.html', mensaje_exito=mensaje_exito, login_session=login_session)
    except Exception as e:
        print(e)
        mensaje_error = "Hubo un error al crear el evento"
        return render_template('eventos.html', mensaje_error=mensaje_error, login_session=login_session)
    

@app.route('/add-question', methods=['POST'])
def add_question():
    try:
        question_text = request.form['question_text']
        answers = request.form['answers'].split(',')
        correct_answer = request.form['correct_answer']
        exam_name = request.form['exam_name']
        table_name = 'Questions_Answers'
        table = dynamodb.Table(table_name)

        item = {
            'question_id': generate_unique_id(), 
            'question_text': question_text,
            'answers': answers,
            'correct_answer': correct_answer,
            'examName': exam_name,
            'user_id': login_session['user_id']
        }
        try:
            response = table.put_item(Item=item)
            print("Pregunta añadida a DynamoDB:", response)
        except NoCredentialsError:
            print("No se encontraron credenciales de AWS")
        return jsonify({'message': 'Pregunta añadida a la base de datos exitosamente'})
    except Exception as e:
        return jsonify({'error': 'Error al añadir la pregunta a la base de datos'})
    
    

@app.route('/obtener_eventos_calendar', methods=['GET'])
def obtener_eventos_calendar():
    try:
        table_name = 'Eventos'
        table = dynamodb.Table(table_name)
        user_id = login_session.get('user_id')
        if user_id is None:
            return jsonify({'error': 'No hay un usuario que haya iniciado sesión'})
        response = table.scan(
            FilterExpression=Attr('user_id').eq(user_id)
        )
        response = table.scan()
        print(response)
        items = response['Items']
        eventos = []
        for item in items:
            start = item.get('fecha_evento', '')
            start = datetime.strptime(start, '%d/%m/%Y').strftime('%d %B, %Y')
            creationDate = item.get('fecha_creacion', '')
            creationDate = datetime.strptime(creationDate, '%d/%m/%Y').strftime('%d %B, %Y')

            evento = {
                'title': item.get('titulo', ''),
                'start': start,
                'description': item.get('descripcion', ''),
                'creationDate': creationDate
            }
            print(evento)
            if evento not in eventos:
                eventos.append(evento)
        print(eventos)
        # Convierte la lista de eventos en una cadena JSON
        return jsonify(eventos)
    except Exception as e:
        print(str(e))
        return jsonify({'error': 'Error leyendo eventos'})


@app.route('/upload', methods=['POST'])
def upload():
    if check_document_limit(login_session):
        return jsonify({'error': 'Has alcanzado tu límite diario de documentos.'}), 400
    
    exam_name = request.form.get('test_name')
    document = request.files['document']
    document_type = str(request.form.get('document-type'))
    table_name = 'Questions_Answers'

    filename = os.path.join(app.config['UPLOAD_FOLDER'], exam_name + document_type)
    document.save(filename)
    table = dynamodb.Table(table_name)

    try:
        if document_type == '.docx':
            extracted_data = extract_questions_answers_from_word(filename, exam_name)
            for item in extracted_data:
                table.put_item(Item=item)
            message = f"Extracted from {document.filename} and stored in DynamoDB with test name: {exam_name}"
        elif document_type == '.xlsx':
            extracted_data = extract_questions_answers_from_excel(filename, exam_name)
            for item in extracted_data:
                table.put_item(Item=item)
            message = f"Extraído de {document.filename} y guardado en la base de datos con nombre de examen: {exam_name}"
        else: 
            raise ValueError("El archivo no es de tipo Word (.docx) o Excel (.xlsx).")
        os.remove(filename)
    except Exception as e:
        message = f"Ocurrió un error con el documento {document.filename}"
    return jsonify({'message': message})


@app.route('/obtener_eventos', methods=['GET'])
def obtener_eventos():
    try:
        table_name = 'Eventos'
        table = dynamodb.Table(table_name)
        response = table.scan()
        eventos = []
        items = response['Items']

        for item in items:
            evento = {}  
            evento_id = item['evento_id']
            tipo = item['tipo']
            estado_evento = item['estado_evento']
            fecha_evento = item['fecha_evento']
            fecha_creacion = item['fecha_creacion']
            titulo = item['titulo']
            descripcion = item['descripcion']
            evento['evento_id'] = evento_id
            evento['tipo'] = tipo
            evento['estado_evento'] = estado_evento
            evento['fecha_evento'] = fecha_evento
            evento['fecha_creacion'] = fecha_creacion
            evento['titulo'] = titulo
            evento['descripcion'] = descripcion
            print(evento)
            eventos.append(evento)
        return jsonify(eventos)
    except Exception as e:
        return jsonify({'error': 'Error leyendo los eventos'})


def extract_questions_answers_from_excel(filename, exam_name):
    results = []
    workbook = load_workbook(filename)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        if len(results) >= MAX_QUESTIONS:
            break
        question_text = row[0]
        possible_answers = int(row[1])
        answers = []
        correct_answer = row[2+possible_answers]
        for indice_columna in range(2, 2+possible_answers):
            answers.append(row[indice_columna])  
        item = {
            'question_id': generate_unique_id(),
            'question_text': question_text,
            'answers': answers,
            'correct_answer': correct_answer,
            'examName': exam_name,
            'user_id': login_session['user_id']
        }
        results.append(item)
    return results




def extract_questions_answers_from_word(docx_file, exam_name):
    results = []
    doc = docx.Document(docx_file)
    current_question = None
    current_answer = []
    correct_answer = None
    for paragraph in doc.paragraphs:
        if len(results) >= MAX_QUESTIONS:
            break
        if re.match(r"^\d+\.", paragraph.text) or (paragraph.runs and paragraph.runs[0].bold):
            if current_question is not None:
                item = {
                    'question_id': generate_unique_id(),
                    'question_text': current_question,
                    'answers': current_answer,
                    'correct_answer': correct_answer,
                    'examName': exam_name,
                    'user_id': login_session['user_id']
                }
                results.append(item)
            current_question = paragraph.text
            current_answer = []
            correct_answer = None
        else:
            match = re.match(r"^[\w\W]+[\.\)]", paragraph.text)
            if match:
                answer_text = paragraph.text[match.end():].strip()
                if paragraph.runs and (paragraph.runs[0].bold or paragraph.runs[0].italic) or "*" in answer_text:
                    if "*" in answer_text:
                        correct_answer = answer_text.replace("*", "").strip() 
                    else:
                        correct_answer = answer_text
                current_answer.append(answer_text)
    if current_question is not None:
        item = {
            'question_id': generate_unique_id(),
            'question_text': current_question,
            'answers': current_answer,
            'correct_answer': correct_answer,
            'examName': exam_name,
            'user_id': login_session['user_id']
        }
        results.append(item)
    return results


def import_cards_from_txt(file, nombre_mazo):
    with open(file, 'r') as file:
        lines = file.read().splitlines()
        items = []
        start_index = 0
        for i, line in enumerate(lines):
            if not line.startswith('#'):
                start_index = i
                break
        for line in lines[start_index:]:
            if len(items) >= MAX_QUESTIONS:
                break
            row = line.split('\t')
            if len(row) < 2:
                continue
            card_id = generate_unique_id()
            item={
                "card_id": card_id,
                "question": row[1],
                "answer": row[0],
                "deck_name":nombre_mazo,
                'user_id': login_session['user_id']
            }
            items.append(item)
        return items



def generate_unique_id():
    return str(uuid.uuid4())

def obtener_eventos_dynamo():
    try:
        table_name = 'Eventos'
        response = dynamodb.scan(TableName=table_name)

        # Obtiene los datos de respuesta
        items = response['Items']
        # Procesa los datos y crea una lista de diccionarios
        eventos = []
        for item in items:
            evento = {
                'titulo': item.get('titulo', {}).get('S', ''),
                'tipo': item.get('tipo', {}).get('S', ''),
                'fecha_creacion': item.get('fecha_creacion', {}).get('S', ''),
                'fecha_evento': item.get('fecha_evento', {}).get('S', ''),
                'descripcion': item.get('descripcion', {}).get('S', ''),
                'estado_evento': item.get('estado_evento', {}).get('S', '')
            }
            eventos.append(evento)
        print(jsonify(eventos))
        return jsonify(eventos)
    except Exception as e:
        return jsonify({'error': 'Error leyendo eventos'})


def obtener_cartas_dynamo(deck_name):
    try:
        table_name = 'CartasParaEstudiar'
        table = dynamodb.Table(table_name) 
        response = table.scan( 
            FilterExpression=Attr('deck_name').eq(deck_name)
        )
        items = response['Items']
        flashcards = []
        for item in items:
            flashcard = {
                'question': item['question'],
                'answer': item['answer'],
            }
            flashcards.append(flashcard)
        return flashcards
    except ClientError as e:
        return jsonify({'error': 'Error leyendo flashcards', 'details': str(e)})
    

def check_document_limit(login_session):
    #Límite de 10 documentos a suir
    doc_limit = 10 
    #Bloque if para comprobar si el usuario ya ha accedido a su límite diario
    if 'doc_count' in login_session and login_session['doc_count'] >= doc_limit:
        return True
    #Se incrementa el recuento de documentos del usuario
    login_session['doc_count'] = login_session.get('doc_count', 0) + 1
    return False


if __name__ == '__main__':
    app.run(debug=True)
